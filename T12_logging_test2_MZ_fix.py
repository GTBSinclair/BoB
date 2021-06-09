####
# Script: Gio S
# Comments: Mat Z
# NOT FOR FLIGHT USE - UNIT TESTING ONLY
####

import smbus
import time
import board
import adafruit_bme680
import openpyxl, schedule, datetime, time

#################
# BME680 SENSOR #
#################

i2c = board.I2C()

## MZ: First changes - following docs for this library
## Docs: https://circuitpython.readthedocs.io/projects/bme680/en/latest/api.html
##sensor = adafruit_bme680.Adafruit_BME680_I2C(i2c, 0x76, False, 2)	# No debug output, 2 readings per second (ensures change).
sensor = adafruit_bme680.Adafruit_BME680_I2C(i2c, 0x76, refresh_rate = 2) # GS. Only takes two positional arguments
sensor.sea_level_pressure = 1013.25		# Assumed for testing


#################
# MS5637 SENSOR #
#################

# Get I2C bus - MZ: Should only need to do this once unless it crashes
bus = smbus.SMBus(1)
addr = 0x77		# Good to keep parameters like this

# MS5637_02BA03 address, 0x76(118)
#		0x1E(30)	Reset command
bus.write_byte(addr, 0x1E)

time.sleep(0.02) 	## MZ: Datasheet says 8.1 ms highest response time. Let's go with 20ms for very high safety. Still 10x less wasted time.

## CALIBRATION SEQUENCE
## MZ: Assuming we only need to calibrate once.

# Read 12 bytes of calibration data
# Read pressure sensitivity
data = bus.read_i2c_block_data(addr, 0xA2, 2)
C1 = data[0] * 256 + data[1]

# Read pressure offset
data = bus.read_i2c_block_data(addr, 0xA4, 2)
C2 = data[0] * 256 + data[1]

# Read temperature coefficient of pressure sensitivity
data = bus.read_i2c_block_data(addr, 0xA6, 2)
C3 = data[0] * 256 + data[1]

# Read temperature coefficient of pressure offset
data = bus.read_i2c_block_data(addr, 0xA8, 2)
C4 = data[0] * 256 + data[1]

# Read reference temperature
data = bus.read_i2c_block_data(addr, 0xAA, 2)
C5 = data[0] * 256 + data[1]

# Read temperature coefficient of the temperature
data = bus.read_i2c_block_data(addr, 0xAC, 2)
C6 = data[0] * 256 + data[1]


def read_MS5637_Data():

	# MS5637_02BA03 address, 0x77(118)
	#		0x40(64)	Pressure conversion(OSR = 256) command
	bus.write_byte(addr, 0x40)

	time.sleep(0.02)

	# Read digital pressure value
	# Read data back from 0x00(0), 3 bytes
	# D1 MSB2, D1 MSB1, D1 LSB
	value = bus.read_i2c_block_data(addr, 0x00, 3)
	D1 = value[0] * 65536 + value[1] * 256 + value[2]

	# MS5637_02BA03 address, 0x76(118)
	#		0x50(64)	Temperature conversion(OSR = 256) command
	bus.write_byte(addr, 0x50)

	time.sleep(0.02)

	# Read digital temperature value
	# Read data back from 0x00(0), 3 bytes
	# D2 MSB2, D2 MSB1, D2 LSB
	value = bus.read_i2c_block_data(addr, 0x00, 3)
	D2 = value[0] * 65536 + value[1] * 256 + value[2]

	## MZ: This is all great stuff, but for easier teamwork please do drop a comment on areas like this referring to the datasheet page where these are specified.
	dT = D2 - C5 * 256
	TEMP = 2000 + dT * C6 / 8388608
	OFF = C2 * 131072 + (C4 * dT) / 64
	SENS = C1 * 65536 + (C3 * dT ) / 128
	T2 = 0
	OFF2 = 0
	SENS2 = 0

	if TEMP > 2000 :
		T2 = 5 * dT * dT / 274877906944
		OFF2 = 0
		SENS2 = 0
	elif TEMP < 2000 :
		T2 = 3 * (dT * dT) / 8589934592
		OFF2 = 61 * ((TEMP - 2000) * (TEMP - 2000)) / 16
		SENS2 = 29 * ((TEMP - 2000) * (TEMP - 2000)) / 16
		if TEMP < -1500:
			OFF2 = OFF2 + 17 * ((TEMP + 1500) * (TEMP + 1500))	## MZ: Fix. Changed to TEMP.
			SENS2 = SENS2 + 9 * ((TEMP + 1500) * (TEMP + 1500))	

	TEMP = TEMP - T2
	OFF = OFF - OFF2
	SENS = SENS - SENS2
	pressure = ((((D1 * SENS) / 2097152) - OFF) / 32768.0) / 100.0
	cTemp = TEMP / 100.0
	fTemp = cTemp * 1.8 + 32

	result = pressure, cTemp, fTemp		# Pack results into a 'tuple'
	return result


####
# MZ: Ok, so everything above seems fine so far. However, consider what is being done. You've made use of the I2C interface to
# write and read various bytes of data to and from the sensor. That's all fine. However you've only saved them once into variables.
# None of the above is written in a repeatable manner - for that you either want to incorporate it within the loop (simplest way),
# or define the above as methods - or for even more structure, as part of a sensor class - which you can repeatedly call within the
# while True loop you define below. See below comment for more info.
####

## MZ: This bit is fine for testing. Not recommended for flight due to memory use, but solid for test outputs. Flight will append to json.
timeStr = str(int(time.time()))
wb = openpyxl.Workbook()
sheet = wb.active
filename = "T12_logging_test_" + timeStr + ".xlsx"
wb.save(filename)

## MZ: Now we only setup the file once
file = openpyxl.load_workbook(filename) 
sheet = file.active
sheet['A1'] = "MS5 Pressure mbar"
sheet['B1'] = "MS5 Temp degC"
sheet['C1'] = "MS5 Temp F"
sheet['E1'] = "BME Temp degC"
sheet['F1'] = "BME Gas"
sheet['G1'] = "BME Humidity %"
sheet['H1'] = "BME Pressure hPa"
sheet['I1'] = "Time s"


## MZ: Your core problem lies within this function below. 
def read_data():

	## MZ: This is fine, you're getting the new time each iteration of the loop.
	x1 = datetime.datetime.now()
	Time = x1.strftime("%d-%m-%y %H:%M:%S")
	print(Time)		## MZ: Would advise against random prints, though, unless you really need real-time visualisation on desktop

	# MZ: Here is where we run our data loop for the other sensor
	pressure, cTemp, fTemp = read_MS5637_Data() 	# MZ: Unpack tuple. We could return the tuple and access using pressure = result[0] for example. Useful tool!

	sheet.cell(column = 1, row = sheet.max_row+1, value = pressure) 	# MZ: max_row is fine for a test case, but poor for production applications. Too many things can go wrong with Excel.
	sheet.cell(column = 2, row = sheet.max_row, value = cTemp)			# MZ: That being said, good job on its use here & the max_row+1 on the first line. Very logical.
	sheet.cell(column = 3, row = sheet.max_row, value = fTemp)

	## MZ: In contrast to the above, these cells work perfectly fine because the class you instantiated for BME680 automatically reads the sensor 10 times a second.
	## Therefore you're updating values automatically without even realising. That is not true for the above.
	sheet.cell(column = 5, row = sheet.max_row, value = sensor.temperature)
	sheet.cell(column = 6, row = sheet.max_row, value = sensor.gas)
	sheet.cell(column = 7, row = sheet.max_row, value = sensor.humidity)
	sheet.cell(column = 8, row = sheet.max_row, value = sensor.pressure)
	sheet.cell(column = 9, row = sheet.max_row, value = Time)

	file.save(filename) 	# MZ: Should work consecutively. If not, check write_only status from https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html 

schedule.every(1).seconds.do(read_data) 	## MZ: You don't really need a library for this, but it's fine for testing and is nice and readable. I'll look into this more later as it is quite clean!

while True:
	schedule.run_pending()






# Output data to screen


#als = True

#while als:

#    print("\n")

#    print("MS5 Pressure : {} mbar".format(pressure))
#    print("MS5 Temperature in Celsius : {} C".format(cTemp))
#    print("MS5 Temperature in Fahrenheit : {} F".format(fTemp))

#    print("\n")

#    print('BME Temperature: {} degrees C'.format(sensor.temperature))
#    print('BME Gas: {} ohms'.format(sensor.gas))
#    print('BME Humidity: {} %'.format(sensor.humidity))
#    print('BME Pressure: {} hPa'.format(sensor.pressure))

#    time.sleep(2)


