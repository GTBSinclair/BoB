import smbus
import time
import board
import adafruit_bme680
import openpyxl, schedule, datetime, time


i2c = board.I2C()
sensor = adafruit_bme680.Adafruit_BME680_I2C(i2c,0x76)

#print("\n")


# Get I2C bus
bus = smbus.SMBus(1)

# MS5637_02BA03 address, 0x76(118)
#		0x1E(30)	Reset command
bus.write_byte(0x77, 0x1E)

time.sleep(0.2)

# Read 12 bytes of calibration data
# Read pressure sensitivity
data = bus.read_i2c_block_data(0x77, 0xA2, 2)
C1 = data[0] * 256 + data[1]

# Read pressure offset
data = bus.read_i2c_block_data(0x77, 0xA4, 2)
C2 = data[0] * 256 + data[1]

# Read temperature coefficient of pressure sensitivity
data = bus.read_i2c_block_data(0x77, 0xA6, 2)
C3 = data[0] * 256 + data[1]

# Read temperature coefficient of pressure offset
data = bus.read_i2c_block_data(0x77, 0xA8, 2)
C4 = data[0] * 256 + data[1]

# Read reference temperature
data = bus.read_i2c_block_data(0x77, 0xAA, 2)
C5 = data[0] * 256 + data[1]

# Read temperature coefficient of the temperature
data = bus.read_i2c_block_data(0x77, 0xAC, 2)
C6 = data[0] * 256 + data[1]

# MS5637_02BA03 address, 0x77(118)
#		0x40(64)	Pressure conversion(OSR = 256) command
bus.write_byte(0x77, 0x40)

time.sleep(0.2)

# Read digital pressure value
# Read data back from 0x00(0), 3 bytes
# D1 MSB2, D1 MSB1, D1 LSB
value = bus.read_i2c_block_data(0x77, 0x00, 3)
D1 = value[0] * 65536 + value[1] * 256 + value[2]

# MS5637_02BA03 address, 0x76(118)
#		0x50(64)	Temperature conversion(OSR = 256) command
bus.write_byte(0x77, 0x50)

time.sleep(0.2)

# Read digital temperature value
# Read data back from 0x00(0), 3 bytes
# D2 MSB2, D2 MSB1, D2 LSB
value = bus.read_i2c_block_data(0x77, 0x00, 3)
D2 = value[0] * 65536 + value[1] * 256 + value[2]

dT = D2 - C5 * 256
TEMP = 2000 + dT * C6 / 8388608
OFF = C2 * 131072 + (C4 * dT) / 64
SENS = C1 * 65536 + (C3 * dT ) / 128
T2 = 0
OFF2 = 0
SENS2 = 0

if TEMP > 2000 :
	T2 = 5 * dT * dT / 274877906944
	OFF2 = 0
	SENS2 = 0
elif TEMP < 2000 :
	T2 = 3 * (dT * dT) / 8589934592
	OFF2 = 61 * ((TEMP - 2000) * (TEMP - 2000)) / 16
	SENS2 = 29 * ((TEMP - 2000) * (TEMP - 2000)) / 16
	if TEMP < -1500:
		OFF2 = OFF2 + 17 * ((Temp + 1500) * (Temp + 1500))
		SENS2 = SENS2 + 9 * ((Temp + 1500) * (Temp +1500))

TEMP = TEMP - T2
OFF = OFF - OFF2
SENS = SENS - SENS2
pressure = ((((D1 * SENS) / 2097152) - OFF) / 32768.0) / 100.0
cTemp = TEMP / 100.0
fTemp = cTemp * 1.8 + 32




wb = openpyxl.Workbook()
sheet = wb.active
wb.save("T12_logging_test.xlsx")


def read_data():

	x1 = datetime.datetime.now()
	Time = x1.strftime("%d-%m-%y %H:%M:%S")
	print(Time)

	file = openpyxl.load_workbook("T12_logging_test.xlsx")
	sheet = file.active

	sheet['A1'] = "MS5 Pressure mbar"
	sheet['B1'] = "MS5 Temp degC"
	sheet['C1'] = "MS5 Temp F"
	sheet['E1'] = "BME Temp degC"
	sheet['F1'] = "BME Gas"
	sheet['G1'] = "BME Humidity %"
	sheet['H1'] = "BME Pressure hPa"
	sheet['I1'] = "Time s"
	sheet.cell(column = 1, row = sheet.max_row+1, value = pressure)
	sheet.cell(column = 2, row = sheet.max_row, value = cTemp)
	sheet.cell(column = 3, row = sheet.max_row, value = fTemp)

	sheet.cell(column = 5, row = sheet.max_row, value = sensor.temperature)
	sheet.cell(column = 6, row = sheet.max_row, value = sensor.gas)
	sheet.cell(column = 7, row = sheet.max_row, value = sensor.humidity)
	sheet.cell(column = 8, row = sheet.max_row, value = sensor.pressure)
	sheet.cell(column = 9, row = sheet.max_row, value = Time)

	file.save("T12_logging_test.xlsx")

schedule.every(1).seconds.do(read_data)

while True:
	schedule.run_pending()


# Output data to screen


#als = True

#while als:

#    print("\n")

#    print("MS5 Pressure : {} mbar".format(pressure))
#    print("MS5 Temperature in Celsius : {} C".format(cTemp))
#    print("MS5 Temperature in Fahrenheit : {} F".format(fTemp))

#    print("\n")

#    print('BME Temperature: {} degrees C'.format(sensor.temperature))
#    print('BME Gas: {} ohms'.format(sensor.gas))
#    print('BME Humidity: {} %'.format(sensor.humidity))
#    print('BME Pressure: {} hPa'.format(sensor.pressure))

#    time.sleep(2)


