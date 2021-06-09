import openpyxl, schedule, datetime, time

wb = openpyxl.Workbook()
sheet = wb.active


wb.save("openpyxl_test.xlsx")


def read_data():

	x1 = datetime.datetime.now()
	Time = x1.strftime("%d-%m-%y %H:%M:%S")
	print(Time)

	file = openpyxl.load_workbook("openpyxl_test.xlsx")

	sheet = file.active

	sheet.cell(column = 1, row = sheet.max_row+1, value = Time)
	file.save("openpyxl_test.xlsx")

schedule.every(2).seconds.do(read_data)

while True:
	schedule.run_pending()
	time.sleep(1)






